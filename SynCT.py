import csv
import logging
import math
import os
import pathlib
import shlex
import shutil
import subprocess
import sys
import tempfile

import vtk

import ctk
import qt
import slicer
from slicer.ScriptedLoadableModule import *
from slicer.util import VTKObservationMixin


#
# SynCT
#


class SynCT(ScriptedLoadableModule):
    """PET/CT preprocessing, registration, and quantitative PET analysis."""

    def __init__(self, parent):
        ScriptedLoadableModule.__init__(self, parent)
        self.parent.title = "SynCT"
        self.parent.categories = ["PET Quantification"]
        self.parent.dependencies = []
        self.parent.contributors = ["SynCT contributors"]
        self.parent.helpText = """
        SynCT provides a compact workflow for PET/CT preprocessing, registration,
        SUVR image generation, ROI statistics, Dice evaluation, and batch PET analysis.
        """
        self.parent.acknowledgementText = "Developed as a 3D Slicer scripted module."


#
# Widget
#


class SynCTWidget(ScriptedLoadableModuleWidget, VTKObservationMixin):
    """Builds a Slicer-native UI without relying on the old .ui file."""

    def __init__(self, parent=None):
        ScriptedLoadableModuleWidget.__init__(self, parent)
        VTKObservationMixin.__init__(self)
        self.logic = None
        self.defaultOutputDir = os.path.join(os.path.dirname(__file__), "tmp_data")

    def setup(self):
        ScriptedLoadableModuleWidget.setup(self)
        self.logic = SynCTLogic()
        os.makedirs(self.defaultOutputDir, exist_ok=True)

        self._buildHeader()
        self._buildDicomSection()
        self._buildPreprocessSection()
        self._buildRegistrationSection()
        self._buildQuantificationSection()
        self._buildGaainAV45Section()
        self._buildBatchSection()
        self._buildResultsSection()
        self.layout.addStretch(1)

    # ------------------------------------------------------------------
    # UI helpers

    def _buildHeader(self):
        self.statusLabel = qt.QLabel("Ready")
        self.progressBar = qt.QProgressBar()
        self.progressBar.setRange(0, 100)
        self.progressBar.setValue(0)

        header = qt.QWidget()
        headerLayout = qt.QHBoxLayout(header)
        headerLayout.setContentsMargins(0, 0, 0, 0)
        headerLayout.addWidget(self.statusLabel, 2)
        headerLayout.addWidget(self.progressBar, 1)
        self.layout.addWidget(header)

    def _section(self, title, collapsed=False):
        section = ctk.ctkCollapsibleButton()
        section.text = title
        section.collapsed = collapsed
        self.layout.addWidget(section)
        layout = qt.QFormLayout(section)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)
        return section, layout

    def _lineEdit(self, text=""):
        line = qt.QLineEdit()
        line.setText(text)
        return line

    def _pathInput(self, mode="open", caption="Select path", filters="All files (*)"):
        edit = qt.QLineEdit()
        button = qt.QPushButton("...")
        button.setMaximumWidth(32)
        row = qt.QWidget()
        rowLayout = qt.QHBoxLayout(row)
        rowLayout.setContentsMargins(0, 0, 0, 0)
        rowLayout.setSpacing(4)
        rowLayout.addWidget(edit)
        rowLayout.addWidget(button)

        def browse():
            start = self._text(edit) or self.defaultOutputDir
            if mode == "dir":
                selected = qt.QFileDialog.getExistingDirectory(slicer.util.mainWindow(), caption, start)
            elif mode == "save":
                selected = qt.QFileDialog.getSaveFileName(slicer.util.mainWindow(), caption, start, filters)
            else:
                selected = qt.QFileDialog.getOpenFileName(slicer.util.mainWindow(), caption, start, filters)
            if isinstance(selected, tuple):
                selected = selected[0]
            if selected:
                edit.setText(str(selected))

        button.connect("clicked(bool)", lambda checked=False: browse())
        return edit, row

    def _nodeSelector(self, nodeTypes=("vtkMRMLScalarVolumeNode",), noneEnabled=True):
        selector = slicer.qMRMLNodeComboBox()
        selector.nodeTypes = list(nodeTypes)
        selector.selectNodeUponCreation = True
        selector.addEnabled = False
        selector.removeEnabled = False
        selector.noneEnabled = noneEnabled
        selector.showHidden = False
        selector.showChildNodeTypes = False
        selector.setMRMLScene(slicer.mrmlScene)
        return selector

    def _combo(self, values, current=None):
        combo = qt.QComboBox()
        for value in values:
            combo.addItem(value)
        if current is not None and current in values:
            combo.setCurrentIndex(values.index(current))
        return combo

    def _button(self, text, callback, tooltip=""):
        button = qt.QPushButton(text)
        button.setToolTip(tooltip)
        button.connect("clicked(bool)", lambda checked=False: callback())
        return button

    def _firstExistingPath(self, candidates):
        for candidate in candidates:
            if candidate and os.path.exists(candidate):
                return candidate
        return ""

    def _defaultTestPatientDir(self):
        testRoot = os.path.join(os.path.dirname(__file__), "test_data")
        if not os.path.isdir(testRoot):
            return ""
        subdirs = [
            os.path.join(testRoot, name)
            for name in os.listdir(testRoot)
            if os.path.isdir(os.path.join(testRoot, name))
        ]
        subdirs.sort()
        return subdirs[0] if subdirs else ""

    def _text(self, widget):
        value = widget.text
        if callable(value):
            value = value()
        return str(value).strip()

    def _comboText(self, combo):
        value = combo.currentText
        if callable(value):
            value = value()
        return str(value)

    def _value(self, widget):
        value = widget.value
        if callable(value):
            value = value()
        return value

    def _checked(self, widget):
        try:
            return bool(widget.isChecked())
        except TypeError:
            return bool(widget.checked)

    def _requiredNode(self, selector, name):
        node = selector.currentNode()
        if node is None:
            raise ValueError(f"Please select {name}.")
        return node

    def _optionalPath(self, edit):
        path = self._text(edit)
        return path if path else None

    def report(self, progress, message):
        self.progressBar.setValue(max(0, min(100, int(progress))))
        self.statusLabel.setText(message)
        if hasattr(self, "logBox"):
            self.logBox.appendPlainText(message)
        logging.info(message)
        slicer.app.processEvents()

    def _run(self, title, callback):
        self.report(0, title)
        try:
            result = callback()
            self.report(100, f"{title}: done")
            return result
        except Exception as exc:
            logging.exception(title)
            self.report(100, f"{title}: failed")
            slicer.util.errorDisplay(str(exc), windowTitle=title)
            return None

    # ------------------------------------------------------------------
    # Sections

    def _buildDicomSection(self):
        _, form = self._section("1. DICOM import")

        self.ctDicomDirEdit, row = self._pathInput("dir", "Select CT DICOM directory")
        form.addRow("CT DICOM dir", row)
        self.petDicomDirEdit, row = self._pathInput("dir", "Select PET DICOM directory")
        form.addRow("PET DICOM dir", row)
        self.dicomOutputDirEdit, row = self._pathInput("dir", "Select output directory")
        self.dicomOutputDirEdit.setText(self.defaultOutputDir)
        form.addRow("Output dir", row)

        self.ctNameEdit = self._lineEdit("CT")
        form.addRow("CT output name", self.ctNameEdit)
        self.petNameEdit = self._lineEdit("PET_SUV")
        form.addRow("PET output name", self.petNameEdit)
        self.petSuvCheck = qt.QCheckBox("Convert PET to SUV using DICOM dose metadata")
        self.petSuvCheck.setChecked(True)
        form.addRow("", self.petSuvCheck)

        row = qt.QWidget()
        rowLayout = qt.QHBoxLayout(row)
        rowLayout.setContentsMargins(0, 0, 0, 0)
        rowLayout.addWidget(self._button("Convert CT", self.onConvertCT))
        rowLayout.addWidget(self._button("Convert PET", self.onConvertPET))
        form.addRow("", row)

    def _buildPreprocessSection(self):
        _, form = self._section("2. Preprocess")

        self.skullInputSelector = self._nodeSelector()
        form.addRow("Skull strip input", self.skullInputSelector)
        self.skullOutputNameEdit = self._lineEdit("brain")
        form.addRow("Brain output", self.skullOutputNameEdit)
        self.skullMaskNameEdit = self._lineEdit("brain_mask")
        form.addRow("Mask output", self.skullMaskNameEdit)
        form.addRow("", self._button("Run skull strip", self.onSkullStrip))

        self.ctClipInputSelector = self._nodeSelector()
        form.addRow("CT clip input", self.ctClipInputSelector)
        self.ctMinSpin = qt.QDoubleSpinBox()
        self.ctMinSpin.setRange(-5000, 5000)
        self.ctMinSpin.setDecimals(1)
        self.ctMinSpin.setValue(-1000.0)
        form.addRow("Clip minimum", self.ctMinSpin)
        self.ctMaxSpin = qt.QDoubleSpinBox()
        self.ctMaxSpin.setRange(-5000, 5000)
        self.ctMaxSpin.setDecimals(1)
        self.ctMaxSpin.setValue(2000.0)
        form.addRow("Clip maximum", self.ctMaxSpin)
        self.ctNormalizeCheck = qt.QCheckBox("Normalize clipped CT to 0-1")
        form.addRow("", self.ctNormalizeCheck)
        self.ctClipNameEdit = self._lineEdit("CT_clipped")
        form.addRow("CT clip output", self.ctClipNameEdit)
        self.ctClipSaveEdit, row = self._pathInput("save", "Save clipped CT", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("Optional save path", row)
        form.addRow("", self._button("Clip CT", self.onClipCT))

        self.maskPetSelector = self._nodeSelector()
        form.addRow("PET volume", self.maskPetSelector)
        self.maskSelector = self._nodeSelector()
        form.addRow("Mask volume", self.maskSelector)
        self.maskedPetNameEdit = self._lineEdit("PET_masked")
        form.addRow("Masked PET output", self.maskedPetNameEdit)
        self.maskedPetSaveEdit, row = self._pathInput("save", "Save masked PET", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("Optional save path", row)
        form.addRow("", self._button("Apply mask to PET", self.onMaskPET))

    def _buildRegistrationSection(self):
        _, form = self._section("3. Registration")

        self.fixedSelector = self._nodeSelector()
        form.addRow("Fixed/reference", self.fixedSelector)
        self.movingSelector = self._nodeSelector()
        form.addRow("Moving/input", self.movingSelector)
        self.rigidOutputNameEdit = self._lineEdit("registered")
        form.addRow("Registered output", self.rigidOutputNameEdit)
        self.rigidTransformNameEdit = self._lineEdit("rigid_transform")
        form.addRow("Transform output", self.rigidTransformNameEdit)
        self.rigidInterpCombo = self._combo(["Linear", "NearestNeighbor", "BSpline", "WindowedSinc"], "BSpline")
        form.addRow("Interpolation", self.rigidInterpCombo)
        self.rigidSamplingSpin = qt.QDoubleSpinBox()
        self.rigidSamplingSpin.setRange(0.001, 1.0)
        self.rigidSamplingSpin.setDecimals(3)
        self.rigidSamplingSpin.setSingleStep(0.01)
        self.rigidSamplingSpin.setValue(0.01)
        form.addRow("Sampling fraction", self.rigidSamplingSpin)
        form.addRow("", self._button("Run rigid registration", self.onRigidRegistration))

        self.applyInputSelector = self._nodeSelector()
        form.addRow("Apply input volume", self.applyInputSelector)
        self.applyTransformSelector = self._nodeSelector(("vtkMRMLTransformNode",))
        form.addRow("Transform", self.applyTransformSelector)
        self.applyReferenceSelector = self._nodeSelector()
        form.addRow("Reference volume", self.applyReferenceSelector)
        self.applyOutputNameEdit = self._lineEdit("transformed")
        form.addRow("Apply output", self.applyOutputNameEdit)
        self.applySaveEdit, row = self._pathInput("save", "Save transformed volume", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("Optional output save path", row)
        self.applyInterpCombo = self._combo(["Linear", "NearestNeighbor", "BSpline", "WindowedSinc"], "BSpline")
        form.addRow("Apply interpolation", self.applyInterpCombo)
        form.addRow("", self._button("Apply transform", self.onApplyTransform))

        self.synthMovingEdit, row = self._pathInput("open", "Select moving image", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("SynthMorph moving", row)
        self.synthFixedEdit, row = self._pathInput("open", "Select fixed image", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("SynthMorph fixed", row)
        self.synthOutputEdit, row = self._pathInput("save", "Save registered image", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("SynthMorph output", row)
        self.synthSaveTransformCheck = qt.QCheckBox("Save deformation/transform field")
        self.synthSaveTransformCheck.setChecked(True)
        form.addRow("", self.synthSaveTransformCheck)
        self.synthTransformEdit, row = self._pathInput("save", "Save transform", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("SynthMorph transform", row)
        self.synthModeCombo = self._combo(["joint", "deform", "affine"], "joint")
        form.addRow("SynthMorph mode", self.synthModeCombo)
        self.synthUseGpuCheck = qt.QCheckBox("Use GPU for SynthMorph registration")
        self.synthUseGpuCheck.setChecked(True)
        form.addRow("", self.synthUseGpuCheck)
        self.synthBackendCombo = self._combo(["Slicer Python", "WSL GPU"], "WSL GPU")
        form.addRow("SynthMorph backend", self.synthBackendCombo)
        self.synthWslDistroEdit = self._lineEdit("Ubuntu")
        self.synthWslDistroEdit.placeholderText = "WSL distro name, e.g. Ubuntu"
        form.addRow("WSL distro", self.synthWslDistroEdit)
        self.synthWslPythonEdit = self._lineEdit("$HOME/envs/synthmorph-gpu/bin/python")
        form.addRow("WSL Python", self.synthWslPythonEdit)
        self.synthCudaEdit = self._lineEdit("0")
        form.addRow("CUDA_VISIBLE_DEVICES", self.synthCudaEdit)
        form.addRow("", self._button("Check SynthMorph backend", self.onCheckSynthBackend))
        form.addRow("", self._button("Run SynthMorph", self.onSynthMorphRegister))

        self.synthApplyTransformEdit, row = self._pathInput(
            "open", "Select SynthMorph transform", "NIfTI/H5 (*.nii *.nii.gz *.h5);;All files (*)"
        )
        form.addRow("SynthMorph apply transform", row)
        self.synthApplyInputEdit, row = self._pathInput(
            "open", "Select image to transform", "NIfTI (*.nii *.nii.gz);;All files (*)"
        )
        form.addRow("SynthMorph apply input", row)
        self.synthApplyOutputEdit, row = self._pathInput(
            "save", "Save transformed image", "NIfTI (*.nii *.nii.gz);;All files (*)"
        )
        form.addRow("SynthMorph apply output", row)
        self.synthApplyInterpCombo = self._combo(["linear", "nearest", "bspline"], "bspline")
        form.addRow("SynthMorph interpolation", self.synthApplyInterpCombo)
        self.synthApplyBackendCombo = self._combo(["Slicer Python", "WSL GPU"], "WSL GPU")
        form.addRow("SynthMorph apply backend", self.synthApplyBackendCombo)
        form.addRow("", self._button("Apply SynthMorph transform", self.onSynthMorphApply))

    def _buildQuantificationSection(self):
        _, form = self._section("4. PET quantification")

        self.suvrPetSelector = self._nodeSelector()
        form.addRow("PET/SUV input", self.suvrPetSelector)
        self.suvrReferenceSelector = self._nodeSelector()
        form.addRow("Reference mask", self.suvrReferenceSelector)
        self.suvrOutputNameEdit = self._lineEdit("SUVR")
        form.addRow("SUVR output", self.suvrOutputNameEdit)
        self.suvrSavePathEdit, row = self._pathInput("save", "Save SUVR image", "NIfTI (*.nii *.nii.gz);;All files (*)")
        form.addRow("Optional save path", row)
        form.addRow("", self._button("Create SUVR image", self.onCreateSUVR))

        self.roiImageSelector = self._nodeSelector()
        form.addRow("ROI intensity image", self.roiImageSelector)
        self.roiLabelSelector = self._nodeSelector()
        form.addRow("ROI label volume", self.roiLabelSelector)
        self.roiLabelsEdit = self._lineEdit("")
        self.roiLabelsEdit.placeholderText = "Blank = all non-zero labels; supports 1,2,5-8"
        form.addRow("ROI labels", self.roiLabelsEdit)
        self.roiTablePathEdit, row = self._pathInput("save", "Save ROI table", "CSV (*.csv);;Excel (*.xlsx);;All files (*)")
        form.addRow("Optional table path", row)
        form.addRow("", self._button("Compute ROI statistics", self.onROIStats))

        self.diceASelector = self._nodeSelector()
        form.addRow("Dice label A", self.diceASelector)
        self.diceBSelector = self._nodeSelector()
        form.addRow("Dice label B", self.diceBSelector)
        self.diceLabelsEdit = self._lineEdit("")
        self.diceLabelsEdit.placeholderText = "Blank = all non-zero labels in A"
        form.addRow("Dice labels", self.diceLabelsEdit)
        self.diceTablePathEdit, row = self._pathInput("save", "Save Dice table", "CSV (*.csv);;Excel (*.xlsx);;All files (*)")
        form.addRow("Optional table path", row)
        form.addRow("", self._button("Compute Dice", self.onDice))

    def _buildBatchSection(self):
        _, form = self._section("6. Batch SUVR + ROI", collapsed=True)

        self.batchRootEdit, row = self._pathInput("dir", "Select subject root directory")
        form.addRow("Subject root", row)
        self.batchPetNameEdit = self._lineEdit("PET_SUV.nii.gz")
        form.addRow("PET filename", self.batchPetNameEdit)
        self.batchLabelNameEdit = self._lineEdit("labels.nii.gz")
        form.addRow("Label filename", self.batchLabelNameEdit)
        self.batchReferenceNameEdit = self._lineEdit("reference_mask.nii.gz")
        form.addRow("Reference mask filename", self.batchReferenceNameEdit)
        self.batchCreateSUVRCheck = qt.QCheckBox("Create SUVR image before ROI statistics")
        self.batchCreateSUVRCheck.setChecked(True)
        form.addRow("", self.batchCreateSUVRCheck)
        self.batchSuvrNameEdit = self._lineEdit("SUVR.nii.gz")
        form.addRow("SUVR output filename", self.batchSuvrNameEdit)
        self.batchLabelsEdit = self._lineEdit("")
        self.batchLabelsEdit.placeholderText = "Blank = all non-zero labels per subject"
        form.addRow("ROI labels", self.batchLabelsEdit)
        self.batchTablePathEdit, row = self._pathInput("save", "Save batch table", "CSV (*.csv);;Excel (*.xlsx);;All files (*)")
        form.addRow("Batch table path", row)
        form.addRow("", self._button("Run batch", self.onBatchSUVRROI))

    def _buildGaainAV45Section(self):
        _, form = self._section("5. GAAIN AV45 SUVR")

        patientDir = self._defaultTestPatientDir()
        self.gaainPatientDirEdit, row = self._pathInput("dir", "Select patient directory")
        self.gaainPatientDirEdit.setText(patientDir)
        form.addRow("Patient directory", row)

        self.gaainCtNameEdit = self._lineEdit("CT.nii")
        form.addRow("CT filename", self.gaainCtNameEdit)
        self.gaainPetNameEdit = self._lineEdit("AV45_PET.nii")
        form.addRow("AV45 PET filename", self.gaainPetNameEdit)

        self.gaainTemplateEdit, row = self._pathInput("open", "Select GAAIN/template image", "NIfTI (*.nii *.nii.gz);;All files (*)")
        self.gaainTemplateEdit.setText(
            self._firstExistingPath(
                [
                    r"D:\hkj\data\data\rspm152.nii.gz",
                    r"D:\hkj\test_data\test_data\003\rspm152.nii.gz",
                    os.path.join(os.path.dirname(__file__), "mri_synthmorph", "data", "rigid", "rbrain_CT.nii.gz"),
                ]
            )
        )
        form.addRow("Template image", row)

        self.gaainCerebellumEdit, row = self._pathInput("open", "Select GAAIN cerebellum reference mask", "NIfTI (*.nii *.nii.gz);;All files (*)")
        self.gaainCerebellumEdit.setText(
            self._firstExistingPath(
                [
                    os.path.join(os.path.dirname(__file__), "Atlas", "Abeta_ctx_meta_roi_2mm", "voi_CerebGry_2mm.nii"),
                    r"D:\hkj\SUVr_mapping_data\Desikan-Killiany_MNI_cerebellumGM.nii",
                    r"D:\hkj\SUVr_mapping_data\1\Desikan-Killiany_MNI_cerebellumGM.nii",
                ]
            )
        )
        form.addRow("GAAIN cerebellum mask", row)

        self.gaainCtxEdit, row = self._pathInput("open", "Select ctx/cortex label mask", "NIfTI (*.nii *.nii.gz);;All files (*)")
        self.gaainCtxEdit.setText(
            self._firstExistingPath(
                [
                    os.path.join(os.path.dirname(__file__), "Atlas", "Abeta_ctx_meta_roi_2mm", "voi_ctx_2mm.nii"),
                    r"D:\hkj\data\data\001\rCortex_label.nii.gz",
                    r"D:\hkj\test_data\test_data\003\rCortex_label.nii.gz",
                    r"D:\hkj\test_data\test_data\003\Cortex_label.nii.gz",
                ]
            )
        )
        form.addRow("ctx label mask", row)

        self.gaainCtxLabelsEdit = self._lineEdit("1")
        form.addRow("ctx labels", self.gaainCtxLabelsEdit)

        defaultOutput = (
            os.path.join(self.defaultOutputDir, "GAAIN_AV45")
            if (not patientDir or any(ord(char) > 127 for char in patientDir))
            else os.path.join(patientDir, "SynCT_GAAIN_AV45")
        )
        self.gaainOutputDirEdit, row = self._pathInput("dir", "Select GAAIN workflow output directory")
        self.gaainOutputDirEdit.setText(defaultOutput)
        form.addRow("Output directory", row)

        self.gaainMniPetNameEdit = self._lineEdit("AV45_PET_MNI_BSpline.nii.gz")
        form.addRow("MNI PET output filename", self.gaainMniPetNameEdit)

        self.gaainRegistrationCombo = self._combo(["joint", "affine", "rigid", "deform"], "joint")
        form.addRow("CT to MNI SynthMorph model", self.gaainRegistrationCombo)
        self.gaainUseGpuCheck = qt.QCheckBox("Use GPU for CT-to-MNI SynthMorph")
        self.gaainUseGpuCheck.setChecked(True)
        form.addRow("", self.gaainUseGpuCheck)
        self.gaainSynthBackendCombo = self._combo(["Slicer Python", "WSL GPU"], "WSL GPU")
        form.addRow("SynthMorph backend", self.gaainSynthBackendCombo)
        self.gaainWslDistroEdit = self._lineEdit("Ubuntu")
        self.gaainWslDistroEdit.placeholderText = "WSL distro name, e.g. Ubuntu"
        form.addRow("WSL distro", self.gaainWslDistroEdit)
        self.gaainWslPythonEdit = self._lineEdit("$HOME/envs/synthmorph-gpu/bin/python")
        form.addRow("WSL Python", self.gaainWslPythonEdit)
        self.gaainCudaEdit = self._lineEdit("0")
        form.addRow("CUDA_VISIBLE_DEVICES", self.gaainCudaEdit)
        form.addRow("", self._button("Check GAAIN SynthMorph backend", self.onCheckGaainSynthBackend))
        self.gaainSamplingSpin = qt.QDoubleSpinBox()
        self.gaainSamplingSpin.setRange(0.001, 1.0)
        self.gaainSamplingSpin.setDecimals(3)
        self.gaainSamplingSpin.setSingleStep(0.01)
        self.gaainSamplingSpin.setValue(0.02)
        form.addRow("PET-CT sampling fraction", self.gaainSamplingSpin)

        self.gaainSaveDeformationCheck = qt.QCheckBox("Save deformation/transform field")
        self.gaainSaveDeformationCheck.setChecked(False)
        form.addRow("", self.gaainSaveDeformationCheck)
        self.gaainDeformationNameEdit = self._lineEdit("CT_to_MNI_deformation.nii.gz")
        form.addRow("Transform field filename", self.gaainDeformationNameEdit)

        interpolationLabel = qt.QLabel("BSpline for CT/PET intensity resampling; nearest-neighbor for masks")
        form.addRow("Interpolation", interpolationLabel)
        form.addRow("", self._button("Run GAAIN AV45 SUVR", self.onGaainAV45Workflow))

    def _buildResultsSection(self):
        _, form = self._section("Results and log")

        self.resultsTable = qt.QTableWidget()
        self.resultsTable.setMinimumHeight(180)
        form.addRow(self.resultsTable)

        self.logBox = qt.QPlainTextEdit()
        self.logBox.setReadOnly(True)
        self.logBox.setMaximumHeight(120)
        form.addRow(self.logBox)

    # ------------------------------------------------------------------
    # Callbacks

    def onConvertCT(self):
        def task():
            result = self.logic.convertDicomToNifti(
                self._text(self.ctDicomDirEdit),
                self._text(self.dicomOutputDirEdit),
                self._text(self.ctNameEdit),
                modality="CT",
                convertPetToSUV=False,
                progressCallback=self.report,
            )
            if result.get("node"):
                self.ctClipInputSelector.setCurrentNode(result["node"])
                self.skullInputSelector.setCurrentNode(result["node"])
            return result

        self._run("Convert CT DICOM", task)

    def onConvertPET(self):
        def task():
            result = self.logic.convertDicomToNifti(
                self._text(self.petDicomDirEdit),
                self._text(self.dicomOutputDirEdit),
                self._text(self.petNameEdit),
                modality="PET",
                convertPetToSUV=self._checked(self.petSuvCheck),
                progressCallback=self.report,
            )
            if result.get("node"):
                self.maskPetSelector.setCurrentNode(result["node"])
                self.suvrPetSelector.setCurrentNode(result["node"])
                self.roiImageSelector.setCurrentNode(result["node"])
            return result

        self._run("Convert PET DICOM", task)

    def onSkullStrip(self):
        def task():
            brain, mask = self.logic.skullStrip(
                self._requiredNode(self.skullInputSelector, "skull strip input"),
                self._text(self.skullOutputNameEdit),
                self._text(self.skullMaskNameEdit),
                progressCallback=self.report,
            )
            self.maskSelector.setCurrentNode(mask)
            self.suvrReferenceSelector.setCurrentNode(mask)
            return {"brain": brain.GetName(), "mask": mask.GetName()}

        self._run("Skull strip", task)

    def onClipCT(self):
        def task():
            node = self.logic.clipVolume(
                self._requiredNode(self.ctClipInputSelector, "CT clip input"),
                float(self._value(self.ctMinSpin)),
                float(self._value(self.ctMaxSpin)),
                self._checked(self.ctNormalizeCheck),
                self._text(self.ctClipNameEdit),
                savePath=self._optionalPath(self.ctClipSaveEdit),
            )
            self.ctClipInputSelector.setCurrentNode(node)
            return {"output": node.GetName()}

        self._run("Clip CT", task)

    def onMaskPET(self):
        def task():
            node = self.logic.maskVolume(
                self._requiredNode(self.maskPetSelector, "PET volume"),
                self._requiredNode(self.maskSelector, "mask volume"),
                self._text(self.maskedPetNameEdit),
                savePath=self._optionalPath(self.maskedPetSaveEdit),
                progressCallback=self.report,
            )
            self.suvrPetSelector.setCurrentNode(node)
            self.roiImageSelector.setCurrentNode(node)
            return {"output": node.GetName()}

        self._run("Apply mask to PET", task)

    def onRigidRegistration(self):
        def task():
            output, transform = self.logic.rigidRegistration(
                fixedNode=self._requiredNode(self.fixedSelector, "fixed/reference volume"),
                movingNode=self._requiredNode(self.movingSelector, "moving/input volume"),
                outputName=self._text(self.rigidOutputNameEdit),
                transformName=self._text(self.rigidTransformNameEdit),
                interpolationMode=self._comboText(self.rigidInterpCombo),
                samplingPercentage=float(self._value(self.rigidSamplingSpin)),
                progressCallback=self.report,
            )
            self.applyInputSelector.setCurrentNode(output)
            self.applyTransformSelector.setCurrentNode(transform)
            return {"output": output.GetName(), "transform": transform.GetName()}

        self._run("Rigid registration", task)

    def onApplyTransform(self):
        def task():
            node = self.logic.applyTransform(
                inputNode=self._requiredNode(self.applyInputSelector, "apply input volume"),
                transformNode=self._requiredNode(self.applyTransformSelector, "transform"),
                referenceNode=self._requiredNode(self.applyReferenceSelector, "reference volume"),
                outputName=self._text(self.applyOutputNameEdit),
                interpolationMode=self._comboText(self.applyInterpCombo),
                savePath=self._optionalPath(self.applySaveEdit),
                progressCallback=self.report,
            )
            self.roiImageSelector.setCurrentNode(node)
            return {"output": node.GetName()}

        self._run("Apply transform", task)

    def onSynthMorphRegister(self):
        def task():
            return self.logic.synthMorphRegister(
                movingPath=self._text(self.synthMovingEdit),
                fixedPath=self._text(self.synthFixedEdit),
                outputPath=self._text(self.synthOutputEdit),
                transformPath=self._text(self.synthTransformEdit),
                saveTransform=self._checked(self.synthSaveTransformCheck),
                mode=self._comboText(self.synthModeCombo),
                useGpu=self._checked(self.synthUseGpuCheck),
                backend=self._comboText(self.synthBackendCombo),
                wslDistro=self._text(self.synthWslDistroEdit),
                wslPython=self._text(self.synthWslPythonEdit),
                cudaVisibleDevices=self._text(self.synthCudaEdit),
                progressCallback=self.report,
            )

        self._run("SynthMorph registration", task)

    def onCheckSynthBackend(self):
        def task():
            result = self.logic.checkSynthMorphBackend(
                backend=self._comboText(self.synthBackendCombo),
                wslDistro=self._text(self.synthWslDistroEdit),
                wslPython=self._text(self.synthWslPythonEdit),
                cudaVisibleDevices=self._text(self.synthCudaEdit),
            )
            self._showTable([result])
            return result

        self._run("Check SynthMorph backend", task)

    def onSynthMorphApply(self):
        def task():
            return self.logic.synthMorphApply(
                transformPath=self._text(self.synthApplyTransformEdit),
                inputPath=self._text(self.synthApplyInputEdit),
                outputPath=self._text(self.synthApplyOutputEdit),
                interpolationMode=self._comboText(self.synthApplyInterpCombo),
                backend=self._comboText(self.synthApplyBackendCombo),
                wslDistro=self._text(self.synthWslDistroEdit),
                wslPython=self._text(self.synthWslPythonEdit),
                cudaVisibleDevices=self._text(self.synthCudaEdit),
                progressCallback=self.report,
            )

        self._run("SynthMorph apply", task)

    def onCreateSUVR(self):
        def task():
            node, info = self.logic.createSUVRImage(
                petNode=self._requiredNode(self.suvrPetSelector, "PET/SUV input"),
                referenceMaskNode=self._requiredNode(self.suvrReferenceSelector, "reference mask"),
                outputName=self._text(self.suvrOutputNameEdit),
                savePath=self._optionalPath(self.suvrSavePathEdit),
                progressCallback=self.report,
            )
            self.roiImageSelector.setCurrentNode(node)
            self._showTable([info])
            return info

        self._run("Create SUVR image", task)

    def onROIStats(self):
        def task():
            imageNode = self._requiredNode(self.roiImageSelector, "ROI intensity image")
            labelNode = self._requiredNode(self.roiLabelSelector, "ROI label volume")
            labels = self.logic.parseLabels(self._text(self.roiLabelsEdit), labelNode=labelNode)
            rows = self.logic.computeROIStatistics(
                imageNode, labelNode, labels, progressCallback=self.report
            )
            saved = self.logic.writeTable(rows, self._optionalPath(self.roiTablePathEdit))
            self._showTable(rows)
            return {"rows": len(rows), "saved": saved}

        self._run("Compute ROI statistics", task)

    def onDice(self):
        def task():
            labelA = self._requiredNode(self.diceASelector, "Dice label A")
            labelB = self._requiredNode(self.diceBSelector, "Dice label B")
            labels = self.logic.parseLabels(self._text(self.diceLabelsEdit), labelNode=labelA)
            rows = self.logic.computeDice(labelA, labelB, labels, progressCallback=self.report)
            saved = self.logic.writeTable(rows, self._optionalPath(self.diceTablePathEdit))
            self._showTable(rows)
            return {"rows": len(rows), "saved": saved}

        self._run("Compute Dice", task)

    def onBatchSUVRROI(self):
        def task():
            labels = self.logic.parseLabels(self._text(self.batchLabelsEdit), labelNode=None)
            rows = self.logic.batchSUVRAndROI(
                rootDir=self._text(self.batchRootEdit),
                petFilename=self._text(self.batchPetNameEdit),
                labelFilename=self._text(self.batchLabelNameEdit),
                referenceMaskFilename=self._text(self.batchReferenceNameEdit),
                createSUVR=self._checked(self.batchCreateSUVRCheck),
                suvrFilename=self._text(self.batchSuvrNameEdit),
                labels=labels,
                tablePath=self._text(self.batchTablePathEdit),
                progressCallback=self.report,
            )
            self._showTable(rows)
            return {"rows": len(rows)}

        self._run("Batch SUVR + ROI", task)

    def onCheckGaainSynthBackend(self):
        def task():
            result = self.logic.checkSynthMorphBackend(
                backend=self._comboText(self.gaainSynthBackendCombo),
                wslDistro=self._text(self.gaainWslDistroEdit),
                wslPython=self._text(self.gaainWslPythonEdit),
                cudaVisibleDevices=self._text(self.gaainCudaEdit),
            )
            self._showTable([result])
            return result

        self._run("Check GAAIN SynthMorph backend", task)

    def onGaainAV45Workflow(self):
        def task():
            labels = self.logic.parseLabels(self._text(self.gaainCtxLabelsEdit), labelNode=None)
            rows = self.logic.runGaainAV45Workflow(
                patientDir=self._text(self.gaainPatientDirEdit),
                ctFilename=self._text(self.gaainCtNameEdit),
                petFilename=self._text(self.gaainPetNameEdit),
                templatePath=self._text(self.gaainTemplateEdit),
                cerebellumMaskPath=self._text(self.gaainCerebellumEdit),
                ctxMaskPath=self._text(self.gaainCtxEdit),
                ctxLabels=labels,
                outputDir=self._text(self.gaainOutputDirEdit),
                mniPetFilename=self._text(self.gaainMniPetNameEdit),
                registrationMode=self._comboText(self.gaainRegistrationCombo),
                samplingPercentage=float(self._value(self.gaainSamplingSpin)),
                useGpu=self._checked(self.gaainUseGpuCheck),
                synthBackend=self._comboText(self.gaainSynthBackendCombo),
                wslDistro=self._text(self.gaainWslDistroEdit),
                wslPython=self._text(self.gaainWslPythonEdit),
                cudaVisibleDevices=self._text(self.gaainCudaEdit),
                saveDeformationField=self._checked(self.gaainSaveDeformationCheck),
                deformationFieldFilename=self._text(self.gaainDeformationNameEdit),
                progressCallback=self.report,
            )
            self._showTable(rows)
            return {"rows": len(rows)}

        self._run("GAAIN AV45 SUVR", task)

    def _showTable(self, rows):
        self.resultsTable.clear()
        if not rows:
            self.resultsTable.setRowCount(0)
            self.resultsTable.setColumnCount(0)
            return

        headers = list(rows[0].keys())
        self.resultsTable.setColumnCount(len(headers))
        self.resultsTable.setHorizontalHeaderLabels(headers)
        self.resultsTable.setRowCount(len(rows))
        for rowIndex, row in enumerate(rows):
            for colIndex, header in enumerate(headers):
                value = row.get(header, "")
                if isinstance(value, float):
                    if math.isfinite(value):
                        value = f"{value:.6g}"
                    else:
                        value = ""
                self.resultsTable.setItem(rowIndex, colIndex, qt.QTableWidgetItem(str(value)))
        self.resultsTable.resizeColumnsToContents()


#
# Logic
#


class SynCTLogic(ScriptedLoadableModuleLogic):
    """Computational backend usable from the Slicer Python console."""

    def __init__(self):
        ScriptedLoadableModuleLogic.__init__(self)

    # ------------------------------------------------------------------
    # General helpers

    def _progress(self, callback, progress, message):
        if callback:
            callback(progress, message)

    def _requirePath(self, path, kind="file"):
        if not path:
            raise ValueError(f"Please provide a {kind} path.")
        pathObj = pathlib.Path(path)
        if kind == "directory":
            if not pathObj.is_dir():
                raise ValueError(f"Directory does not exist: {pathObj}")
        elif not pathObj.is_file():
            raise ValueError(f"File does not exist: {pathObj}")
        return pathObj

    def _ensureOutputFile(self, path):
        pathObj = pathlib.Path(path)
        if pathObj.parent:
            pathObj.parent.mkdir(parents=True, exist_ok=True)
        return pathObj

    def _requireOutputPath(self, path, description="output"):
        if not str(path or "").strip():
            raise ValueError(f"Please provide a {description} path.")
        return self._ensureOutputFile(path)

    def _nodeName(self, name, fallback):
        name = (name or "").strip()
        if not name:
            name = fallback
        return slicer.mrmlScene.GetUniqueNameByString(name)

    def _niftiStem(self, value, fallback):
        name = pathlib.Path(str(value or fallback)).name
        lower = name.lower()
        if lower.endswith(".nii.gz"):
            return name[:-7]
        if lower.endswith(".nii"):
            return name[:-4]
        return pathlib.Path(name).stem

    def _copyGeometry(self, sourceNode, targetNode):
        if hasattr(targetNode, "CopyOrientation"):
            targetNode.CopyOrientation(sourceNode)
            return
        matrix = vtk.vtkMatrix4x4()
        sourceNode.GetIJKToRASMatrix(matrix)
        targetNode.SetIJKToRASMatrix(matrix)

    def _newVolumeLike(self, sourceNode, array, name):
        outputNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLScalarVolumeNode", self._nodeName(name, "SynCTVolume")
        )
        slicer.util.updateVolumeFromArray(outputNode, array)
        self._copyGeometry(sourceNode, outputNode)
        outputNode.CreateDefaultDisplayNodes()
        return outputNode

    def _saveNode(self, node, savePath):
        if not savePath:
            return None
        outputPath = self._ensureOutputFile(savePath)
        if not slicer.util.saveNode(node, str(outputPath)):
            raise RuntimeError(f"Failed to save node to: {outputPath}")
        return str(outputPath)

    def _nodeIJKToRASAffine(self, node):
        np = self._np()
        matrix = vtk.vtkMatrix4x4()
        node.GetIJKToRASMatrix(matrix)
        return np.array(
            [[matrix.GetElement(row, col) for col in range(4)] for row in range(4)],
            dtype="float64",
        )

    def _storageFilePath(self, node):
        storageNode = node.GetStorageNode() if node else None
        if not storageNode:
            return None
        try:
            path = storageNode.GetFullNameFromFileName()
        except AttributeError:
            path = storageNode.GetFileName()
        return path if path and os.path.exists(path) else None

    def _saveArrayLikeNode(self, arrayKJI, referenceNode, savePath):
        if not savePath:
            return None
        np = self._np()
        nib = self._nib()
        outputPath = self._ensureOutputFile(savePath)
        dataIJK = np.asarray(arrayKJI, dtype="float32").transpose(2, 1, 0)
        affine = self._nodeIJKToRASAffine(referenceNode)
        header = None
        referencePath = self._storageFilePath(referenceNode)
        if referencePath:
            try:
                header = nib.load(referencePath).header.copy()
            except Exception:
                header = None
        image = nib.Nifti1Image(dataIJK, affine, header=header)
        image.set_sform(affine, code=1)
        image.set_qform(affine, code=1)
        nib.save(image, str(outputPath))
        return str(outputPath)

    def _showVolume(self, node):
        try:
            slicer.util.setSliceViewerLayers(background=node, fit=True)
        except TypeError:
            slicer.util.setSliceViewerLayers(background=node)
            slicer.util.resetSliceViews()

    def _loadVolumeFromFile(self, path, name=None):
        pathObj = self._requirePath(path)
        nodeName = self._nodeName(name or self._niftiStem(pathObj, pathObj.stem), self._niftiStem(pathObj, "volume"))
        node = slicer.util.loadVolume(str(pathObj), properties={"name": nodeName})
        if not node:
            raise RuntimeError(f"Failed to load volume: {pathObj}")
        return node

    def _module(self, moduleName):
        if not hasattr(slicer.modules, moduleName):
            raise RuntimeError(f"Required Slicer module is not available: {moduleName}")
        return getattr(slicer.modules, moduleName)

    def _cliCompleted(self, cliNode):
        status = cliNode.GetStatusString()
        if status != "Completed":
            raise RuntimeError(f"CLI failed with status: {status}")

    def _geometryMatches(self, nodeA, nodeB, tolerance=1e-4):
        if not nodeA or not nodeB:
            return False
        if not nodeA.GetImageData() or not nodeB.GetImageData():
            return False
        if nodeA.GetImageData().GetDimensions() != nodeB.GetImageData().GetDimensions():
            return False
        matrixA = vtk.vtkMatrix4x4()
        matrixB = vtk.vtkMatrix4x4()
        nodeA.GetIJKToRASMatrix(matrixA)
        nodeB.GetIJKToRASMatrix(matrixB)
        for row in range(4):
            for col in range(4):
                if abs(matrixA.GetElement(row, col) - matrixB.GetElement(row, col)) > tolerance:
                    return False
        return True

    def _resampleToReference(self, inputNode, referenceNode, outputName, interpolationMode="NearestNeighbor"):
        outputNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLScalarVolumeNode", self._nodeName(outputName, "SynCT_resampled")
        )
        parameters = {
            "inputVolume": inputNode.GetID(),
            "referenceVolume": referenceNode.GetID(),
            "outputVolume": outputNode.GetID(),
            "interpolationMode": interpolationMode,
        }
        cliNode = slicer.cli.runSync(self._module("brainsresample"), None, parameters)
        self._cliCompleted(cliNode)
        outputNode.CreateDefaultDisplayNodes()
        return outputNode

    def _matchingLabelNode(self, labelNode, referenceNode, progressCallback=None):
        if self._geometryMatches(labelNode, referenceNode):
            return labelNode, False
        self._progress(progressCallback, 25, "Resampling label/mask to reference geometry")
        return self._resampleToReference(
            labelNode, referenceNode, "SynCT_resampled_label", "NearestNeighbor"
        ), True

    # ------------------------------------------------------------------
    # DICOM conversion

    def convertDicomToNifti(
        self,
        dicomDir,
        outputDir,
        outputName,
        modality="CT",
        convertPetToSUV=False,
        progressCallback=None,
    ):
        dicomDir = self._requirePath(dicomDir, "directory")
        outputDir = pathlib.Path(outputDir or os.path.join(os.path.dirname(__file__), "tmp_data"))
        outputDir.mkdir(parents=True, exist_ok=True)
        cleanName = self._niftiStem(outputName, modality)
        outputPath = outputDir / f"{cleanName}.nii.gz"

        self._progress(progressCallback, 10, f"Converting {modality} DICOM to NIfTI")

        try:
            import dicom2nifti
        except Exception as exc:
            raise RuntimeError("dicom2nifti is required for DICOM import.") from exc

        with tempfile.TemporaryDirectory() as tempDir:
            dicom2nifti.convert_directory(str(dicomDir), tempDir, compression=True, reorient=True)
            niftiFiles = list(pathlib.Path(tempDir).glob("*.nii")) + list(pathlib.Path(tempDir).glob("*.nii.gz"))
            if not niftiFiles:
                raise RuntimeError("No NIfTI file was produced by dicom2nifti.")
            niftiFile = max(niftiFiles, key=lambda p: p.stat().st_size)

            if modality.upper() == "PET" and convertPetToSUV:
                self._progress(progressCallback, 55, "Applying PET SUV conversion")
                suvFactor = self.calculateSUVFactorFromDicomDirectory(dicomDir)
                import nibabel as nib

                petImage = nib.load(str(niftiFile))
                petData = petImage.get_fdata(dtype="float32") * float(suvFactor)
                suvImage = nib.Nifti1Image(petData.astype("float32"), petImage.affine, petImage.header)
                nib.save(suvImage, str(outputPath))
            else:
                shutil.copyfile(str(niftiFile), str(outputPath))

        self._progress(progressCallback, 80, f"Loading {outputPath.name} into Slicer")
        node = slicer.util.loadVolume(str(outputPath), properties={"name": cleanName})
        if not node:
            raise RuntimeError(f"Failed to load converted image: {outputPath}")
        self._showVolume(node)
        return {"path": str(outputPath), "node": node}

    def calculateSUVFactorFromDicomDirectory(self, dicomDir):
        try:
            import pydicom
        except Exception as exc:
            raise RuntimeError("pydicom is required for PET SUV conversion.") from exc

        dicomFile = None
        dataset = None
        for candidate in pathlib.Path(dicomDir).rglob("*"):
            if not candidate.is_file():
                continue
            try:
                dataset = pydicom.dcmread(str(candidate), stop_before_pixels=True)
                dicomFile = candidate
                break
            except Exception:
                continue

        if dataset is None:
            raise RuntimeError(f"No readable DICOM file found in: {dicomDir}")

        try:
            radiopharm = dataset.RadiopharmaceuticalInformationSequence[0]
            totalDose = float(radiopharm.RadionuclideTotalDose)
            halfLife = float(radiopharm.RadionuclideHalfLife)
            startTime = str(radiopharm.RadiopharmaceuticalStartTime)
            acquisitionValue = getattr(dataset, "AcquisitionTime", None) or getattr(dataset, "SeriesTime", None)
            if acquisitionValue is None:
                raise AttributeError("AcquisitionTime or SeriesTime is missing")
            acquisitionTime = str(acquisitionValue)
            patientWeightKg = float(dataset.PatientWeight)
        except Exception as exc:
            raise RuntimeError(
                f"Missing SUV DICOM metadata in {dicomFile}. "
                "Need dose, half-life, injection time, acquisition time, and patient weight."
            ) from exc

        elapsedSeconds = self._dicomTimeToSeconds(acquisitionTime) - self._dicomTimeToSeconds(startTime)
        if elapsedSeconds < 0:
            elapsedSeconds += 24 * 3600
        decayedDose = totalDose * (0.5 ** (elapsedSeconds / halfLife))
        if decayedDose <= 0:
            raise RuntimeError("Invalid decayed dose computed from DICOM metadata.")
        return 1000.0 * patientWeightKg / decayedDose

    def _dicomTimeToSeconds(self, value):
        value = str(value).split("+")[0].split("-")[0].replace(":", "")
        if "." in value:
            main, frac = value.split(".", 1)
            fraction = float("0." + frac)
        else:
            main, fraction = value, 0.0
        main = main.ljust(6, "0")
        return int(main[0:2]) * 3600 + int(main[2:4]) * 60 + int(main[4:6]) + fraction

    # ------------------------------------------------------------------
    # Preprocessing

    def skullStrip(self, inputNode, outputName, maskName, progressCallback=None):
        self._progress(progressCallback, 10, "Starting SwissSkullStripper")
        outputNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLScalarVolumeNode", self._nodeName(outputName, "brain")
        )
        maskNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", self._nodeName(maskName, "brain_mask")
        )
        parameters = {
            "patientVolume": inputNode.GetID(),
            "patientOutputVolume": outputNode.GetID(),
            "patientMaskLabel": maskNode.GetID(),
        }
        cliNode = slicer.cli.runSync(self._module("swissskullstripper"), None, parameters)
        self._cliCompleted(cliNode)
        outputNode.CreateDefaultDisplayNodes()
        maskNode.CreateDefaultDisplayNodes()
        self._showVolume(outputNode)
        return outputNode, maskNode

    def clipVolume(self, inputNode, minimum, maximum, normalize, outputName, savePath=None):
        if maximum <= minimum:
            raise ValueError("Clip maximum must be greater than clip minimum.")
        array = slicer.util.arrayFromVolume(inputNode).astype("float32")
        clipped = array.clip(float(minimum), float(maximum))
        if normalize:
            clipped = (clipped - float(minimum)) / (float(maximum) - float(minimum))
        outputNode = self._newVolumeLike(inputNode, clipped.astype("float32"), outputName or "CT_clipped")
        self._saveNode(outputNode, savePath)
        self._showVolume(outputNode)
        return outputNode

    def maskVolume(self, imageNode, maskNode, outputName, savePath=None, progressCallback=None):
        maskNode, temporary = self._matchingLabelNode(maskNode, imageNode, progressCallback)
        try:
            imageArray = slicer.util.arrayFromVolume(imageNode).astype("float32")
            maskArray = slicer.util.arrayFromVolume(maskNode)
            if imageArray.shape != maskArray.shape:
                raise RuntimeError(f"Shape mismatch after resampling: {imageArray.shape} vs {maskArray.shape}")
            outputArray = imageArray * (maskArray > 0)
            outputNode = self._newVolumeLike(imageNode, outputArray.astype("float32"), outputName or "masked")
            self._saveNode(outputNode, savePath)
            self._showVolume(outputNode)
            return outputNode
        finally:
            if temporary:
                slicer.mrmlScene.RemoveNode(maskNode)

    # ------------------------------------------------------------------
    # Registration

    def rigidRegistration(
        self,
        fixedNode,
        movingNode,
        outputName,
        transformName,
        interpolationMode="Linear",
        samplingPercentage=0.01,
        useAffine=False,
        progressCallback=None,
    ):
        self._progress(progressCallback, 10, "Running BRAINSFit rigid registration")
        outputNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLScalarVolumeNode", self._nodeName(outputName, "registered")
        )
        transformNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLTransformNode", self._nodeName(transformName, "rigid_transform")
        )
        parameters = {
            "fixedVolume": fixedNode.GetID(),
            "movingVolume": movingNode.GetID(),
            "outputVolume": outputNode.GetID(),
            "outputTransform": transformNode.GetID(),
            "useRigid": True,
            "useAffine": bool(useAffine),
            "initializeTransformMode": "useGeometryAlign",
            "samplingPercentage": float(samplingPercentage),
            "interpolationMode": interpolationMode,
        }
        cliNode = slicer.cli.runSync(self._module("brainsfit"), None, parameters)
        self._cliCompleted(cliNode)
        outputNode.CreateDefaultDisplayNodes()
        self._showVolume(outputNode)
        return outputNode, transformNode

    def applyTransform(
        self,
        inputNode,
        transformNode,
        referenceNode,
        outputName,
        interpolationMode="Linear",
        savePath=None,
        progressCallback=None,
    ):
        self._progress(progressCallback, 20, "Applying transform with BRAINSResample")
        outputNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLScalarVolumeNode", self._nodeName(outputName, "transformed")
        )
        parameters = {
            "inputVolume": inputNode.GetID(),
            "referenceVolume": referenceNode.GetID(),
            "outputVolume": outputNode.GetID(),
            "warpTransform": transformNode.GetID(),
            "interpolationMode": interpolationMode,
        }
        cliNode = slicer.cli.runSync(self._module("brainsresample"), None, parameters)
        self._cliCompleted(cliNode)
        outputNode.CreateDefaultDisplayNodes()
        self._saveNode(outputNode, savePath)
        self._showVolume(outputNode)
        return outputNode

    def checkSynthMorphBackend(
        self,
        backend="Slicer Python",
        wslDistro="Ubuntu",
        wslPython="$HOME/envs/synthmorph-gpu/bin/python",
        cudaVisibleDevices="0",
    ):
        probe = (
            "import tensorflow as tf; "
            "print('TensorFlow', tf.__version__); "
            "print('GPUs', tf.config.list_physical_devices('GPU'))"
        )
        if self._synthBackendName(backend) == "wsl":
            command = self._wslCommand(
                [(wslPython or "python3").strip(), "-c", probe],
                distro=wslDistro,
                cudaVisibleDevices=cudaVisibleDevices,
            )
        else:
            command = [self._pythonSlicerExecutable(), "-c", probe]
        result, stdoutText, stderrText = self._runSubprocess(command)
        if result.returncode != 0:
            raise RuntimeError(self._subprocessFailureText(stdoutText, stderrText, "SynthMorph backend check failed."))
        output = (stdoutText or stderrText).strip()
        gpuVisible = "PhysicalDevice" in output and "GPU" in output
        return {
            "Backend": self._synthBackendName(backend),
            "CUDA_VISIBLE_DEVICES": str(cudaVisibleDevices or "") if self._synthBackendName(backend) == "wsl" else "",
            "GPUVisible": bool(gpuVisible),
            "Output": output[-1000:],
        }

    def synthMorphRegister(
        self,
        movingPath,
        fixedPath,
        outputPath,
        transformPath,
        mode="joint",
        saveTransform=True,
        useGpu=True,
        backend="Slicer Python",
        wslDistro="Ubuntu",
        wslPython="$HOME/envs/synthmorph-gpu/bin/python",
        cudaVisibleDevices="0",
        progressCallback=None,
    ):
        movingPath = self._requirePath(movingPath)
        fixedPath = self._requirePath(fixedPath)
        outputPath = self._requireOutputPath(outputPath, "SynthMorph output")
        transformPath = self._requireOutputPath(transformPath, "SynthMorph transform") if saveTransform else None
        scriptPath = pathlib.Path(__file__).parent / "mri_synthmorph" / "mri_synthmorph.py"
        if not scriptPath.is_file():
            raise RuntimeError(f"SynthMorph script not found: {scriptPath}")
        self._progress(progressCallback, 10, f"Starting SynthMorph subprocess ({self._synthBackendName(backend)})")
        command = self._synthMorphCommand(
            [
                str(scriptPath),
                "register",
                str(movingPath),
                str(fixedPath),
                "-m",
                mode,
                "-o",
                str(outputPath),
            ],
            backend=backend,
            wslDistro=wslDistro,
            wslPython=wslPython,
            cudaVisibleDevices=cudaVisibleDevices if useGpu else "",
            pathIndices=(0, 2, 3, 7),
        )
        if transformPath:
            self._appendSynthMorphArgs(
                command,
                ["-t", str(transformPath)],
                backend=backend,
                pathIndices=(1,),
            )
        if useGpu:
            self._appendSynthMorphArgs(command, ["-g"], backend=backend)
        result, stdoutText, stderrText = self._runSubprocess(command)
        if result.returncode != 0:
            raise RuntimeError(self._subprocessFailureText(stdoutText, stderrText, "SynthMorph failed."))
        self._progress(progressCallback, 85, "Loading SynthMorph output")
        node = slicer.util.loadVolume(str(outputPath), properties={"name": self._niftiStem(outputPath, "synthmorph")})
        if node:
            self._showVolume(node)
        return {
            "output": str(outputPath),
            "transform": str(transformPath) if transformPath else "",
            "backend": self._synthBackendName(backend),
            "stdout": stdoutText[-2000:],
        }

    def synthMorphApply(
        self,
        transformPath,
        inputPath,
        outputPath,
        interpolationMode="linear",
        backend="Slicer Python",
        wslDistro="Ubuntu",
        wslPython="$HOME/envs/synthmorph-gpu/bin/python",
        cudaVisibleDevices="0",
        progressCallback=None,
    ):
        transformPath = self._requirePath(transformPath)
        inputPath = self._requirePath(inputPath)
        outputPath = self._requireOutputPath(outputPath, "SynthMorph apply output")
        scriptPath = pathlib.Path(__file__).parent / "mri_synthmorph" / "mri_synthmorph.py"
        if not scriptPath.is_file():
            raise RuntimeError(f"SynthMorph script not found: {scriptPath}")
        interpolationMode = self._synthInterpolationName(interpolationMode)
        self._progress(progressCallback, 10, f"Applying SynthMorph transform ({self._synthBackendName(backend)})")
        command = self._synthMorphCommand(
            [
                str(scriptPath),
                "apply",
                str(transformPath),
                str(inputPath),
                str(outputPath),
                "-m",
                interpolationMode,
            ],
            backend=backend,
            wslDistro=wslDistro,
            wslPython=wslPython,
            cudaVisibleDevices=cudaVisibleDevices,
            pathIndices=(0, 2, 3, 4),
        )
        result, stdoutText, stderrText = self._runSubprocess(command)
        if result.returncode != 0:
            raise RuntimeError(self._subprocessFailureText(stdoutText, stderrText, "SynthMorph apply failed."))
        self._progress(progressCallback, 85, "Loading transformed image")
        node = slicer.util.loadVolume(str(outputPath), properties={"name": self._niftiStem(outputPath, "transformed")})
        if node:
            self._showVolume(node)
        return {
            "output": str(outputPath),
            "backend": self._synthBackendName(backend),
            "stdout": stdoutText[-2000:],
        }

    def _synthMorphCommand(
        self,
        args,
        backend="Slicer Python",
        wslDistro="Ubuntu",
        wslPython="$HOME/envs/synthmorph-gpu/bin/python",
        cudaVisibleDevices="",
        pathIndices=(),
    ):
        backend = self._synthBackendName(backend)
        args = list(args)
        if backend == "wsl":
            converted = [
                self._windowsPathToWslPath(arg) if index in pathIndices else str(arg)
                for index, arg in enumerate(args)
            ]
            pythonExecutable = (wslPython or "python3").strip()
            return self._wslCommand(
                [pythonExecutable] + converted,
                distro=wslDistro,
                cudaVisibleDevices=cudaVisibleDevices,
            )
        return [self._pythonSlicerExecutable()] + [str(arg) for arg in args]

    def _appendSynthMorphArgs(self, command, args, backend="Slicer Python", pathIndices=()):
        if self._synthBackendName(backend) == "wsl":
            shell = command[-1]
            converted = [
                self._windowsPathToWslPath(arg) if index in pathIndices else str(arg)
                for index, arg in enumerate(args)
            ]
            command[-1] = shell + " " + " ".join(self._wslShellToken(arg) for arg in converted)
        else:
            command.extend(str(arg) for arg in args)

    def _synthBackendName(self, backend):
        value = str(backend or "Slicer Python").strip().lower()
        return "wsl" if "wsl" in value else "slicer"

    def _wslCommand(self, args, distro="", cudaVisibleDevices=""):
        command = ["wsl"]
        distro = str(distro or "").strip()
        if distro:
            command.extend(["-d", distro])
        command.extend(["--", "bash", "-lc"])
        exports = []
        if str(cudaVisibleDevices or "").strip():
            exports.append(f"export CUDA_VISIBLE_DEVICES={shlex.quote(str(cudaVisibleDevices).strip())}")
        exports.append("export TF_FORCE_GPU_ALLOW_GROWTH=true")
        shellCommand = "; ".join(exports + [" ".join(self._wslShellToken(arg) for arg in args)])
        command.append(shellCommand)
        return command

    def _wslShellToken(self, value):
        text = str(value)
        if text in ("python", "python3"):
            return text
        if text.startswith("$HOME/") or text.startswith("${HOME}/") or text.startswith("~/"):
            return text
        return shlex.quote(text)

    def _windowsPathToWslPath(self, path):
        text = str(path)
        if not text:
            return text
        normalizedInput = text.replace("\\", "/")
        if normalizedInput.startswith("/"):
            return normalizedInput
        pathObj = pathlib.Path(text)
        try:
            pathObj = pathObj.resolve()
        except OSError:
            pathObj = pathObj.absolute()
        normalized = str(pathObj).replace("\\", "/")
        if len(normalized) >= 2 and normalized[1] == ":":
            drive = normalized[0].lower()
            rest = normalized[2:].lstrip("/")
            return f"/mnt/{drive}/{rest}"
        raise RuntimeError(f"Cannot convert path to WSL path: {text}")

    def _runSubprocess(self, command):
        try:
            result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except FileNotFoundError as exc:
            if command and str(command[0]).lower() == "wsl":
                raise RuntimeError(
                    "WSL was selected as the SynthMorph backend, but wsl.exe was not found. "
                    "Install WSL2 and Ubuntu, then rerun this workflow."
                ) from exc
            raise
        stdoutText = self._decodeProcessOutput(result.stdout)
        stderrText = self._decodeProcessOutput(result.stderr)
        if stdoutText:
            logging.info(stdoutText[-2000:])
        if stderrText:
            logging.warning(stderrText[-2000:])
        return result, stdoutText, stderrText

    def _subprocessFailureText(self, stdoutText, stderrText, fallback):
        texts = [text.strip() for text in (stderrText, stdoutText) if text and text.strip()]
        if not texts:
            return fallback
        informative = [
            text for text in texts
            if "Traceback" in text or "Error" in text or "Exception" in text or len(text) > 20
        ]
        return max(informative or texts, key=len)

    def _decodeProcessOutput(self, data):
        if data is None:
            return ""
        if isinstance(data, str):
            return data
        if not data:
            return ""
        try:
            utf8Text = data.decode("utf-8")
        except UnicodeDecodeError:
            utf8Text = data.decode("utf-8", errors="replace")
        utf8Text = self._cleanProcessText(utf8Text)
        diagnosticTokens = ("Traceback", "Error", "Exception", "WARNING", "Created device")
        if any(token in utf8Text for token in diagnosticTokens):
            return utf8Text
        if data.startswith((b"\xff\xfe", b"\xfe\xff")) or self._looksLikeUtf16(data):
            try:
                return self._cleanProcessText(data.decode("utf-16"))
            except UnicodeDecodeError:
                pass
        for encoding in ("gbk", "mbcs"):
            try:
                return self._cleanProcessText(data.decode(encoding))
            except (UnicodeDecodeError, LookupError):
                continue
        return utf8Text

    def _cleanProcessText(self, text):
        lines = str(text or "").replace("\x00", "").splitlines()
        lines = [
            line for line in lines
            if not (line.startswith("wsl:") and "localhost" in line and "WSL" in line)
        ]
        return "\n".join(lines)

    def _looksLikeUtf16(self, data):
        if len(data) < 8:
            return False
        evenNulls = data[0::2].count(b"\x00")
        oddNulls = data[1::2].count(b"\x00")
        halfLength = max(len(data) // 2, 1)
        dominant = max(evenNulls, oddNulls)
        other = min(evenNulls, oddNulls)
        return dominant / halfLength > 0.2 and other <= max(2, dominant // 4)

    def _synthInterpolationName(self, interpolationMode):
        value = str(interpolationMode or "linear").strip().lower()
        aliases = {
            "nearestneighbor": "nearest",
            "nearest neighbor": "nearest",
            "bspline": "bspline",
            "b-spline": "bspline",
            "b spline": "bspline",
            "linear": "linear",
            "nearest": "nearest",
        }
        return aliases.get(value, "linear")

    def applyTransformFileToVolume(
        self,
        transformPath,
        inputNode,
        referenceNode,
        outputName,
        interpolationMode="BSpline",
        savePath=None,
        fallbackInputPath=None,
        synthBackend="Slicer Python",
        wslDistro="Ubuntu",
        wslPython="$HOME/envs/synthmorph-gpu/bin/python",
        cudaVisibleDevices="0",
        progressCallback=None,
    ):
        transformPath = self._requirePath(transformPath)
        transformNode = None
        synthInterpolation = self._synthInterpolationName(interpolationMode)
        self._lastTransformApplyInterpolation = interpolationMode
        if self._isSynthMorphTransformFile(transformPath):
            if not fallbackInputPath or not savePath:
                raise RuntimeError("SynthMorph transform application requires an input file and output save path.")
            self._progress(progressCallback, 20, "Applying SynthMorph deformation field to PET")
            self._lastTransformApplyInterpolation = f"{synthInterpolation} (SynthMorph apply)"
            result = self.synthMorphApply(
                transformPath=transformPath,
                inputPath=fallbackInputPath,
                outputPath=savePath,
                interpolationMode=synthInterpolation,
                backend=synthBackend,
                wslDistro=wslDistro,
                wslPython=wslPython,
                cudaVisibleDevices=cudaVisibleDevices,
                progressCallback=progressCallback,
            )
            node = slicer.mrmlScene.GetFirstNodeByName(self._niftiStem(result["output"], outputName))
            return node or self._loadVolumeFromFile(result["output"], outputName)
        try:
            self._progress(progressCallback, 20, "Loading deformation field into Slicer")
            transformNode = slicer.util.loadTransform(str(transformPath))
            if not transformNode:
                raise RuntimeError(f"Failed to load transform: {transformPath}")
            return self.applyTransform(
                inputNode=inputNode,
                transformNode=transformNode,
                referenceNode=referenceNode,
                outputName=outputName,
                interpolationMode=interpolationMode,
                savePath=savePath,
                progressCallback=progressCallback,
            )
        except Exception:
            if not fallbackInputPath or not savePath:
                raise
            logging.warning("Could not apply transform with BRAINSResample; falling back to SynthMorph apply.")
            self._lastTransformApplyInterpolation = f"{synthInterpolation} (SynthMorph apply fallback)"
            result = self.synthMorphApply(
                transformPath=transformPath,
                inputPath=fallbackInputPath,
                outputPath=savePath,
                interpolationMode=synthInterpolation,
                backend=synthBackend,
                wslDistro=wslDistro,
                wslPython=wslPython,
                cudaVisibleDevices=cudaVisibleDevices,
                progressCallback=progressCallback,
            )
            node = slicer.mrmlScene.GetFirstNodeByName(self._niftiStem(result["output"], outputName))
            return node or self._loadVolumeFromFile(result["output"], outputName)
        finally:
            if transformNode:
                slicer.mrmlScene.RemoveNode(transformNode)

    def _isSynthMorphTransformFile(self, path):
        lower = str(path).lower()
        return lower.endswith((".nii", ".nii.gz", ".mgz", ".lta"))

    def _pythonSlicerExecutable(self):
        candidates = []
        slicerApp = getattr(slicer, "app", None)
        slicerHome = getattr(slicerApp, "slicerHome", "") if slicerApp else ""
        if callable(slicerHome):
            slicerHome = slicerHome()
        if slicerHome:
            executableName = "PythonSlicer.exe" if os.name == "nt" else "PythonSlicer"
            candidates.append(pathlib.Path(slicerHome) / "bin" / executableName)
        if sys.executable:
            candidates.append(pathlib.Path(sys.executable))
        for candidate in candidates:
            if candidate and candidate.is_file():
                return str(candidate)
        raise RuntimeError("Could not locate PythonSlicer executable.")

    # ------------------------------------------------------------------
    # Quantification

    def createSUVRImage(self, petNode, referenceMaskNode, outputName, savePath=None, progressCallback=None):
        referenceMaskNode, temporary = self._matchingLabelNode(referenceMaskNode, petNode, progressCallback)
        try:
            petArray = slicer.util.arrayFromVolume(petNode).astype("float32")
            maskArray = slicer.util.arrayFromVolume(referenceMaskNode)
            referenceValues = petArray[maskArray > 0]
            referenceValues = referenceValues[~self._np().isnan(referenceValues)]
            if referenceValues.size == 0:
                raise RuntimeError("Reference mask contains no PET voxels.")
            referenceMean = float(referenceValues.mean())
            if abs(referenceMean) < 1e-12:
                raise RuntimeError("Reference mean is zero; cannot compute SUVR.")
            suvrArray = petArray / referenceMean
            outputNode = self._newVolumeLike(petNode, suvrArray.astype("float32"), outputName or "SUVR")
            savedPath = self._saveArrayLikeNode(suvrArray, petNode, savePath) if savePath else None
            self._showVolume(outputNode)
            info = {
                "Metric": "SUVR image",
                "Output": outputNode.GetName(),
                "ReferenceMean": referenceMean,
                "ReferenceVoxelCount": int(referenceValues.size),
                "SavedPath": savedPath or "",
            }
            return outputNode, info
        finally:
            if temporary:
                slicer.mrmlScene.RemoveNode(referenceMaskNode)

    def runGaainAV45Workflow(
        self,
        patientDir,
        ctFilename,
        petFilename,
        templatePath,
        cerebellumMaskPath,
        ctxMaskPath,
        ctxLabels,
        outputDir,
        mniPetFilename="AV45_PET_MNI_BSpline.nii.gz",
        registrationMode="Rigid+Affine",
        samplingPercentage=0.02,
        useGpu=True,
        synthBackend="Slicer Python",
        wslDistro="Ubuntu",
        wslPython="$HOME/envs/synthmorph-gpu/bin/python",
        cudaVisibleDevices="0",
        saveDeformationField=False,
        deformationFieldFilename="CT_to_MNI_deformation.nii.gz",
        progressCallback=None,
    ):
        patientDir = self._requirePath(patientDir, "directory")
        outputDir = pathlib.Path(outputDir or (patientDir / "SynCT_GAAIN_AV45"))
        outputDir.mkdir(parents=True, exist_ok=True)
        subject = patientDir.name

        ctPath = self._requirePath(patientDir / (ctFilename or "CT.nii"))
        petPath = self._requirePath(patientDir / (petFilename or "AV45_PET.nii"))
        templatePath = self._requirePath(templatePath)
        cerebellumMaskPath = self._requirePath(cerebellumMaskPath)
        ctxMaskPath = self._requirePath(ctxMaskPath)

        self._progress(progressCallback, 5, "Loading CT, AV45 PET, and template")
        ctNode = self._loadVolumeFromFile(ctPath, f"{subject}_CT")
        petNode = self._loadVolumeFromFile(petPath, f"{subject}_AV45_PET")
        templateNode = self._loadVolumeFromFile(templatePath, "GAAIN_template")

        synthMorphMode = str(registrationMode or "joint").strip().lower()
        if synthMorphMode not in ("joint", "affine", "rigid", "deform"):
            synthMorphMode = "joint"
        petRigidPath = outputDir / "AV45_PET_rigid_to_native_CT.nii.gz"
        petToCtTransformPath = outputDir / "PET_to_native_CT_rigid_transform.h5"
        ctBrainPath = outputDir / "native_CT_brain.nii.gz"
        ctBrainMaskPath = outputDir / "native_CT_brain_mask.nii.gz"
        petBrainPath = outputDir / "AV45_PET_native_CT_brain.nii.gz"
        ctClipPath = outputDir / "native_CT_brain_HU0_80.nii.gz"
        ctToTemplatePath = outputDir / "native_CT_brain_HU0_80_MNI.nii.gz"
        transformPath = outputDir / (deformationFieldFilename or "CT_to_MNI_deformation.nii.gz")
        if not saveDeformationField:
            transformPath = outputDir / ".SynCT_temp_CT_to_MNI_deformation.nii.gz"
        mniPetPath = outputDir / (mniPetFilename or "AV45_PET_MNI_BSpline.nii.gz")
        cerebellumInPetPath = outputDir / "GAAIN_cerebellumGM_in_AV45_template.nii.gz"
        ctxInPetPath = outputDir / "GAAIN_ctx_in_AV45_template.nii.gz"
        suvrPath = outputDir / "AV45_SUVR_GAAIN_cerebellumGM.nii.gz"
        tablePath = outputDir / "AV45_GAAIN_ctx_SUVR.csv"
        savedTransformPath = ""
        savedPetToCtTransformPath = ""

        try:
            self._progress(progressCallback, 12, "Rigidly registering AV45 PET to native CT with BSpline interpolation")
            petCtNode, petCtTransformNode = self.rigidRegistration(
                fixedNode=ctNode,
                movingNode=petNode,
                outputName="AV45_PET_rigid_to_native_CT",
                transformName="PET_to_native_CT_rigid_transform",
                interpolationMode="BSpline",
                samplingPercentage=samplingPercentage,
                useAffine=False,
                progressCallback=progressCallback,
            )
            self._saveNode(petCtNode, petRigidPath)
            if saveDeformationField:
                savedPetToCtTransformPath = self._saveNode(petCtTransformNode, petToCtTransformPath) or ""

            self._progress(progressCallback, 25, "Skull stripping native CT")
            ctBrainNode, ctBrainMaskNode = self.skullStrip(
                ctNode,
                "native_CT_brain",
                "native_CT_brain_mask",
                progressCallback=progressCallback,
            )
            self._saveNode(ctBrainNode, ctBrainPath)
            self._saveNode(ctBrainMaskNode, ctBrainMaskPath)

            self._progress(progressCallback, 38, "Applying CT brain mask to rigidly aligned AV45 PET")
            petBrainNode = self.maskVolume(
                petCtNode,
                ctBrainMaskNode,
                "AV45_PET_native_CT_brain",
                savePath=petBrainPath,
                progressCallback=progressCallback,
            )

            self._progress(progressCallback, 50, "Clipping skull-stripped CT to HU [0, 80]")
            ctClipNode = self.clipVolume(
                ctBrainNode,
                0.0,
                80.0,
                False,
                "native_CT_brain_HU0_80",
                savePath=ctClipPath,
            )

            self._progress(progressCallback, 62, "Estimating CT-guided deformation from clipped CT to MNI with SynthMorph")
            self.synthMorphRegister(
                movingPath=ctClipPath,
                fixedPath=templatePath,
                outputPath=ctToTemplatePath,
                transformPath=transformPath,
                mode=synthMorphMode,
                saveTransform=True,
                useGpu=useGpu,
                backend=synthBackend,
                wslDistro=wslDistro,
                wslPython=wslPython,
                cudaVisibleDevices=cudaVisibleDevices,
                progressCallback=progressCallback,
            )
            ctMniNode = self._loadVolumeFromFile(ctToTemplatePath, "native_CT_brain_HU0_80_MNI")
            savedTransformPath = str(transformPath) if saveDeformationField else ""

            self._progress(progressCallback, 74, "Applying CT-guided MNI deformation to AV45 PET")
            petTemplateNode = self.applyTransformFileToVolume(
                transformPath=transformPath,
                inputNode=petBrainNode,
                referenceNode=templateNode,
                outputName="AV45_PET_MNI_BSpline",
                interpolationMode="BSpline",
                savePath=mniPetPath,
                fallbackInputPath=petBrainPath,
                synthBackend=synthBackend,
                wslDistro=wslDistro,
                wslPython=wslPython,
                cudaVisibleDevices=cudaVisibleDevices,
                progressCallback=progressCallback,
            )

            self._progress(progressCallback, 82, "Saving GAAIN masks in MNI PET geometry")
            self._saveLabelImageInReferenceFile(cerebellumMaskPath, mniPetPath, cerebellumInPetPath)
            self._saveLabelImageInReferenceFile(ctxMaskPath, mniPetPath, ctxInPetPath)
            cerebellumInPetNode = self._loadVolumeFromFile(
                cerebellumInPetPath,
                self._nodeName("GAAIN_cerebellumGM_in_AV45_template", "GAAIN_cerebellumGM"),
            )
            ctxInPetNode = self._loadVolumeFromFile(
                ctxInPetPath,
                self._nodeName("GAAIN_ctx_in_AV45_template", "GAAIN_ctx"),
            )

            self._progress(progressCallback, 88, "Creating SUVR image directly from MNI PET and cerebellum reference")
            suvrInfo = self.createSUVRImageFileWithInfo(
                mniPetPath,
                cerebellumInPetPath,
                suvrPath,
            )
            suvrNode = self._loadVolumeFromFile(suvrPath, "AV45_SUVR_GAAIN_cerebellumGM")
            self._showVolume(suvrNode)
        finally:
            if not saveDeformationField and transformPath.exists():
                try:
                    transformPath.unlink()
                except OSError:
                    logging.warning("Could not remove temporary deformation field: %s", transformPath)

        self._progress(progressCallback, 92, "Computing ctx SUVR statistics")
        nib = self._nib()
        suvrImage = nib.load(str(suvrPath))
        ctxImage = nib.load(str(ctxInPetPath))
        suvrArray = suvrImage.get_fdata(dtype="float32")
        ctxArray = self._niftiDataInReference(ctxImage, suvrImage, order=0)
        if not ctxLabels:
            ctxLabels = self._labelsFromArray(ctxArray)

        rows = [
            self._combinedSUVRRow(
                suvrArray,
                ctxArray,
                ctxLabels,
                subject=subject,
                region="ctx",
                referenceMean=suvrInfo["ReferenceMean"],
            )
        ]
        for row in self._roiRowsFromArrays(suvrArray, ctxArray, ctxLabels, subject=subject):
            rows.append(self._suvrROIOutputRow(row, suvrInfo["ReferenceMean"]))

        for row in rows:
            row["ReferenceRegion"] = "GAAIN cerebellum GM"
            row["Registration"] = f"PET-to-CT rigid + CT-guided SynthMorph {synthMorphMode}"
            row["SynthMorphBackend"] = self._synthBackendName(synthBackend)
            row["CUDA_VISIBLE_DEVICES"] = str(cudaVisibleDevices or "") if self._synthBackendName(synthBackend) == "wsl" else ""
            row["ImageInterpolation"] = (
                "PET-to-CT: BSpline; CT clip to MNI: SynthMorph; PET CT-to-MNI: "
                f"{getattr(self, '_lastTransformApplyInterpolation', 'BSpline')}"
            )
            row["MaskInterpolation"] = "NearestNeighbor"
            row["SynthMorphGPURequested"] = bool(useGpu)
            row["SUVRImage"] = str(suvrPath)
            row["MNI_PET"] = str(mniPetPath)
            row["MNI_CT"] = str(ctToTemplatePath)
            row["PETRigidToCT"] = str(petRigidPath)
            row["PETBrainNativeCT"] = str(petBrainPath)
            row["CTClipHU0_80"] = str(ctClipPath)
            row["DeformationField"] = savedTransformPath
            row["PETToCTTransform"] = savedPetToCtTransformPath

        self.writeTable(rows, tablePath)
        self._showVolume(suvrNode)
        self._progress(progressCallback, 100, f"GAAIN AV45 ctx SUVR saved: {tablePath}")
        return rows

    def _combinedSUVRRow(self, imageArray, labelArray, labels, subject, region, referenceMean):
        np = self._np()
        mask = np.isin(labelArray, labels)
        values = imageArray[mask]
        values = values[np.isfinite(values)]
        return {
            "Subject": subject,
            "Region": region,
            "Labels": ",".join(str(label) for label in labels),
            "VoxelCount": int(values.size),
            "SUVR_Mean": float(values.mean()) if values.size else float("nan"),
            "SUVR_Std": float(values.std()) if values.size else float("nan"),
            "SUVR_Min": float(values.min()) if values.size else float("nan"),
            "SUVR_Max": float(values.max()) if values.size else float("nan"),
            "ReferenceMean": float(referenceMean),
        }

    def _suvrROIOutputRow(self, row, referenceMean):
        return {
            "Subject": row.get("Subject", ""),
            "Region": f"ctx_label_{row.get('Label')}",
            "Labels": row.get("Label", ""),
            "VoxelCount": row.get("VoxelCount", 0),
            "SUVR_Mean": row.get("Mean", float("nan")),
            "SUVR_Std": row.get("Std", float("nan")),
            "SUVR_Min": row.get("Min", float("nan")),
            "SUVR_Max": row.get("Max", float("nan")),
            "ReferenceMean": float(referenceMean),
        }

    def computeROIStatistics(self, imageNode, labelNode, labels, progressCallback=None):
        labelNode, temporary = self._matchingLabelNode(labelNode, imageNode, progressCallback)
        try:
            imageArray = slicer.util.arrayFromVolume(imageNode).astype("float32")
            labelArray = slicer.util.arrayFromVolume(labelNode)
            return self._roiRowsFromArrays(imageArray, labelArray, labels)
        finally:
            if temporary:
                slicer.mrmlScene.RemoveNode(labelNode)

    def computeDice(self, labelNodeA, labelNodeB, labels, progressCallback=None):
        labelNodeB, temporary = self._matchingLabelNode(labelNodeB, labelNodeA, progressCallback)
        try:
            np = self._np()
            arrayA = slicer.util.arrayFromVolume(labelNodeA)
            arrayB = slicer.util.arrayFromVolume(labelNodeB)
            rows = []
            for label in labels:
                maskA = arrayA == label
                maskB = arrayB == label
                volumeA = int(np.count_nonzero(maskA))
                volumeB = int(np.count_nonzero(maskB))
                denominator = volumeA + volumeB
                dice = float(2.0 * np.count_nonzero(maskA & maskB) / denominator) if denominator else float("nan")
                rows.append(
                    {
                        "Label": label,
                        "Dice": dice,
                        "VoxelCountA": volumeA,
                        "VoxelCountB": volumeB,
                    }
                )
            return rows
        finally:
            if temporary:
                slicer.mrmlScene.RemoveNode(labelNodeB)

    def _roiRowsFromArrays(self, imageArray, labelArray, labels, subject=None):
        np = self._np()
        rows = []
        if not labels:
            labels = self._labelsFromArray(labelArray)
        for label in labels:
            mask = labelArray == label
            values = imageArray[mask]
            values = values[np.isfinite(values)]
            row = {}
            if subject is not None:
                row["Subject"] = subject
            row.update(
                {
                    "Label": label,
                    "VoxelCount": int(values.size),
                    "Mean": float(values.mean()) if values.size else float("nan"),
                    "Std": float(values.std()) if values.size else float("nan"),
                    "Min": float(values.min()) if values.size else float("nan"),
                    "Max": float(values.max()) if values.size else float("nan"),
                }
            )
            rows.append(row)
        return rows

    def parseLabels(self, text, labelNode=None):
        text = (text or "").strip()
        if not text:
            if labelNode is None:
                return []
            return self._labelsFromArray(slicer.util.arrayFromVolume(labelNode))

        labels = []
        for part in text.replace(";", ",").split(","):
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                start, end = part.split("-", 1)
                start, end = int(start.strip()), int(end.strip())
                step = 1 if end >= start else -1
                labels.extend(range(start, end + step, step))
            else:
                labels.append(int(part))
        return labels

    def _labelsFromArray(self, labelArray):
        np = self._np()
        values = np.unique(labelArray)
        values = values[np.isfinite(values)]
        labels = []
        for value in values:
            if value == 0:
                continue
            rounded = int(round(float(value)))
            if abs(float(value) - rounded) < 1e-4:
                labels.append(rounded)
        return labels

    # ------------------------------------------------------------------
    # Batch file workflow

    def batchSUVRAndROI(
        self,
        rootDir,
        petFilename,
        labelFilename,
        referenceMaskFilename,
        createSUVR,
        suvrFilename,
        labels,
        tablePath,
        progressCallback=None,
    ):
        root = self._requirePath(rootDir, "directory")
        if not tablePath:
            raise ValueError("Please provide a batch table output path.")
        subjects = sorted([path for path in root.iterdir() if path.is_dir()])
        if not subjects:
            raise RuntimeError(f"No subject subdirectories found in: {root}")

        allRows = []
        errors = []
        for index, subjectDir in enumerate(subjects):
            progress = int((index / max(len(subjects), 1)) * 95)
            self._progress(progressCallback, progress, f"Batch processing {subjectDir.name}")
            try:
                petPath = subjectDir / petFilename
                labelPath = subjectDir / labelFilename
                if not petPath.is_file():
                    raise FileNotFoundError(f"PET not found: {petPath}")
                if not labelPath.is_file():
                    raise FileNotFoundError(f"Label not found: {labelPath}")

                statImagePath = petPath
                if createSUVR:
                    referencePath = subjectDir / referenceMaskFilename
                    if not referencePath.is_file():
                        raise FileNotFoundError(f"Reference mask not found: {referencePath}")
                    statImagePath = subjectDir / suvrFilename
                    self.createSUVRImageFile(petPath, referencePath, statImagePath)

                rows = self.computeROIStatisticsFile(statImagePath, labelPath, labels, subjectDir.name)
                allRows.extend(rows)
            except Exception as exc:
                errors.append({"Subject": subjectDir.name, "Error": str(exc)})
                logging.exception("Batch failed for %s", subjectDir)

        if errors:
            allRows.extend(errors)
        self.writeTable(allRows, tablePath)
        self._progress(progressCallback, 100, f"Batch table saved: {tablePath}")
        return allRows

    def createSUVRImageFile(self, petPath, referenceMaskPath, outputPath):
        info = self.createSUVRImageFileWithInfo(petPath, referenceMaskPath, outputPath)
        return info["SavedPath"]

    def createSUVRImageFileWithInfo(self, petPath, referenceMaskPath, outputPath):
        np = self._np()
        nib = self._nib()
        petImage = nib.load(str(petPath))
        referenceImage = nib.load(str(referenceMaskPath))
        petData = petImage.get_fdata(dtype="float32")
        referenceData = self._niftiDataInReference(referenceImage, petImage, order=0)
        referenceValues = petData[referenceData > 0]
        referenceValues = referenceValues[np.isfinite(referenceValues)]
        if referenceValues.size == 0:
            raise RuntimeError("Reference mask contains no PET voxels.")
        referenceMean = float(referenceValues.mean())
        if abs(referenceMean) < 1e-12:
            raise RuntimeError("Reference mean is zero; cannot compute SUVR.")
        suvrData = petData / referenceMean
        outputPath = self._ensureOutputFile(outputPath)
        header = petImage.header.copy()
        header.set_data_dtype("float32")
        outputImage = nib.Nifti1Image(suvrData.astype("float32"), petImage.affine, header=header)
        outputImage.set_sform(petImage.affine, code=int(petImage.header["sform_code"]) or 1)
        outputImage.set_qform(petImage.affine, code=int(petImage.header["qform_code"]) or 1)
        nib.save(outputImage, str(outputPath))
        return {
            "Metric": "SUVR image",
            "Output": pathlib.Path(outputPath).stem,
            "ReferenceMean": referenceMean,
            "ReferenceVoxelCount": int(referenceValues.size),
            "SavedPath": str(outputPath),
        }

    def _saveLabelImageInReferenceFile(self, labelPath, referencePath, outputPath):
        np = self._np()
        nib = self._nib()
        labelImage = nib.load(str(labelPath))
        referenceImage = nib.load(str(referencePath))
        labelData = self._niftiDataInReference(labelImage, referenceImage, order=0)
        labelData = np.rint(labelData)
        sourceDtype = labelImage.get_data_dtype()
        if not np.issubdtype(sourceDtype, np.integer):
            sourceDtype = np.int16
        outputPath = self._ensureOutputFile(outputPath)
        header = referenceImage.header.copy()
        header.set_data_dtype(sourceDtype)
        outputImage = nib.Nifti1Image(labelData.astype(sourceDtype), referenceImage.affine, header=header)
        outputImage.set_sform(referenceImage.affine, code=int(referenceImage.header["sform_code"]) or 1)
        outputImage.set_qform(referenceImage.affine, code=int(referenceImage.header["qform_code"]) or 1)
        nib.save(outputImage, str(outputPath))
        return str(outputPath)

    def computeROIStatisticsFile(self, imagePath, labelPath, labels, subject):
        nib = self._nib()
        image = nib.load(str(imagePath))
        label = nib.load(str(labelPath))
        imageData = image.get_fdata(dtype="float32")
        labelData = self._niftiDataInReference(label, image, order=0)
        if not labels:
            labels = self._labelsFromArray(labelData)
        return self._roiRowsFromArrays(imageData, labelData, labels, subject=subject)

    def _niftiDataInReference(self, sourceImage, referenceImage, order=0):
        np = self._np()
        if (
            sourceImage.shape[:3] == referenceImage.shape[:3]
            and np.allclose(sourceImage.affine, referenceImage.affine, atol=1e-4)
        ):
            return sourceImage.get_fdata()

        try:
            from scipy.ndimage import map_coordinates
        except Exception as exc:
            raise RuntimeError("scipy is required to resample file-based NIfTI images.") from exc

        sourceData = sourceImage.get_fdata()
        referenceShape = referenceImage.shape[:3]
        grid = np.meshgrid(
            np.arange(referenceShape[0]),
            np.arange(referenceShape[1]),
            np.arange(referenceShape[2]),
            indexing="ij",
        )
        flatReference = np.vstack([axis.reshape(1, -1) for axis in grid] + [np.ones((1, grid[0].size))])
        sourceCoordinates = np.linalg.inv(sourceImage.affine).dot(referenceImage.affine).dot(flatReference)
        sampled = map_coordinates(
            sourceData,
            [
                sourceCoordinates[0].reshape(referenceShape),
                sourceCoordinates[1].reshape(referenceShape),
                sourceCoordinates[2].reshape(referenceShape),
            ],
            order=order,
            mode="constant",
            cval=0.0,
        )
        return sampled

    # ------------------------------------------------------------------
    # Table writing and lazy imports

    def writeTable(self, rows, outputPath):
        if not outputPath:
            return None
        if not rows:
            raise RuntimeError("No rows to write.")
        outputPath = pathlib.Path(outputPath)
        outputPath.parent.mkdir(parents=True, exist_ok=True)
        headers = []
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)

        if outputPath.suffix.lower() == ".xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "SynCT"
                worksheet.append(headers)
                for row in rows:
                    worksheet.append([row.get(header, "") for header in headers])
                for columnIndex, header in enumerate(headers, start=1):
                    worksheet.column_dimensions[get_column_letter(columnIndex)].width = max(
                        12, min(32, len(header) + 4)
                    )
                workbook.save(str(outputPath))
                return str(outputPath)
            except Exception:
                logging.exception("Failed to write xlsx; falling back to csv.")
                outputPath = outputPath.with_suffix(".csv")

        with outputPath.open("w", newline="", encoding="utf-8-sig") as stream:
            writer = csv.DictWriter(stream, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        return str(outputPath)

    def _np(self):
        import numpy as np

        return np

    def _nib(self):
        try:
            import nibabel as nib
            return nib
        except Exception as exc:
            raise RuntimeError("nibabel is required for file-based NIfTI processing.") from exc


#
# Test
#


class SynCTTest(ScriptedLoadableModuleTest):
    def setUp(self):
        slicer.mrmlScene.Clear(0)

    def runTest(self):
        self.setUp()
        logic = SynCTLogic()
        self.assertEqual(logic.parseLabels("1,2,5-7"), [1, 2, 5, 6, 7])
